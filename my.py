# coding=utf-8
import openpyxl
f  = open('food.txt','r',encoding='UTF-8-sig')
data=f.read().split("{\"recipes\":[")[1].split("],")[0].replace("},{","},,{").split(",,")
f.close()
lists=[]

for i in range(0,len(data)):
    data[i]=eval(data[i])
    if data[i]["got"]=="是":
        lists.append(data[i]["id"])
wb = openpyxl.load_workbook('0807.xlsx')
sh = wb['Sheet1']
for i in range(16,549):
    if sh.cell(i,3).value in lists:
        sh.cell(i,8).value=1
    else:
        sh.cell(i,8).value=0
wb.save('0807.xlsx')
wb.close()
# listtest=[]
# for i in range(16,537):
#     if sh.cell(i,3).value in lists:
#         listtest.append(sh.cell(i,3).value)
# listtest.sort()
# print(listtest)
# print(len(listtest))
# print(sh.cell(16,3).value)
